import sys
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_sample_bcr_tcr(xlsx_path, sheet_name):
    wb = load_workbook(filename=xlsx_path, read_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No sheet with name '{sheet_name}'")

    sheet = wb[sheet_name]

    count_tr = 0
    count_br = 0
    count_plus_tr = 0
    count_plus_br = 0

    for row in sheet.iter_rows(min_row=2, max_row=6):
        v_gene = row[1].value
        d_gene = row[2].value
        j_gene = row[3].value
        igv_check = row[4].value
        if v_gene and isinstance(v_gene, str): 
            if v_gene.strip().startswith('TR'):
                count_tr += 1
                if igv_check.strip()[0] == "+":
                    count_plus_tr += 1
            if v_gene.strip().startswith('IG'):
                count_br += 1
                if igv_check.strip()[0] == "+":
                    count_plus_br += 1
        if d_gene and isinstance(d_gene, str):
            if d_gene.strip().startswith('TR'):
                count_tr += 1
                if igv_check.strip()[1] == "+":
                    count_plus_tr += 1
            if d_gene.strip().startswith('IG'):
                count_br += 1
                if igv_check.strip()[1] == "+":
                    count_plus_br += 1
        if j_gene and isinstance(j_gene, str):
            if j_gene.strip().startswith('TR'):
                count_tr += 1
                if igv_check.strip()[2] == "+":
                    count_plus_tr += 1
            if j_gene.strip().startswith('IG'):
                count_br += 1
                if igv_check.strip()[2] == "+":
                    count_plus_br += 1

    return count_tr, count_br, count_plus_tr, count_plus_br


def parse_xlsx(xlsx_path, sheet_name, tool_col_ind):
    wb = load_workbook(filename=xlsx_path, read_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No sheet with name '{sheet_name}'")

    sheet = wb[sheet_name]
    
    TP = 0
    TN = 0
    FP = 0
    FN = 0
    
    for row in sheet.iter_rows(min_row=2, max_row=6):
        clon_found = row[5].value
        tool_found = row[tool_col_ind].value
        if clon_found == "+" and tool_found == "+":
            TP += 1
        if clon_found == "-" and tool_found == "-":
            TN += 1
        if clon_found == "+" and tool_found == "-":
            FN += 1
        if clon_found == "-" and tool_found == "+":
            FP += 1
            
    return TP, TN, FP, FN
     
     
def precision_calc(TP, FP):
    try:
        res = TP/(TP+FP)
    except:
        return 0
    else:
        return res
    
def recall_calc(TP, FN):
    try:
        res = TP/(TP+FN)
    except:
        return 0
    else:
        return res

def f1_calc(Pr, Recall):
    try:
        res = 2*((Pr*Recall)/(Pr+Recall))
    except:
        return 0
    else:
        return res
    
def specificity_calc(TN, FP):
    try:
        res = TN/(TN+FP)
    except:
        return 0
    else:
        return res
    
def accuracy_calc(TP, TN, FP, FN):
    try:
        res = (TP+TN)/(TP+TN+FP+FN)
    except:
        return 0
    else:
        return res


def create_stat_table_bcr_tcr(samples_clones, cov_file, outfile):
    cov_df = pd.read_csv(cov_file, sep="\t", names=['sample', 'coverage'], dtype={'sample':'string'})
    cov_df = cov_df.set_index('sample')

    res_df = pd.DataFrame(columns=['sample', 'BCR_mixcr', 'TCR_mixcr', 'BCR_igv', 'TCR_igv', 'Precision_TCR', 'Precision_BCR', 'Median Coverage'])
    
    for i in range(30):
        sample = "sample" + str(i+1)
        print(sample)
        try:
            tr, br, plus_tr, plus_br = parse_sample_bcr_tcr(samples_clones, sample)
        except Exception as e:
            print(e)
        else:
            prec_br = pd.NA if br == 0 else plus_br / br
            prec_tr = pd.NA if tr == 0 else plus_tr / tr
            res_df.loc[len(res_df)] = [sample, br, tr, plus_br, plus_tr, prec_br, prec_tr,cov_df.loc[sample, 'coverage']]
            #res_df.loc[len(res_df)] = [sample, br, tr, plus_br, plus_tr, plus_trust_bam,plus_trust_fastq,plus_mixcr,prec_br, prec_tr,prec_trust_bam,prec_trust_fastq,prec_mixcr]
            
    res_df.to_csv(outfile,sep="\t", index=False)
         

def create_stat_table(samples_clones, outfile, tool_col_ind):

    res_df = pd.DataFrame(columns=['sample', 'Precision', 'Recall', 'F1', 'Specificity', 'Accuracy'])

    for i in range(30):
        sample = "sample" + str(i+1)
        print(sample)
        try:
            TP, TN, FP, FN = parse_xlsx(samples_clones, sample, tool_col_ind)
        except Exception as e:
            print(e)
        else:
            Pr = precision_calc(TP, FP)
            Recall = recall_calc(TP, FN)
            res_df.loc[len(res_df)] = [sample, Pr, Recall, f1_calc(Pr, Recall), specificity_calc(TN, FP), accuracy_calc(TP, TN, FP, FN)]

    res_df.to_csv(outfile,sep="\t", index=False)

def concat_tables(file1, file2, file3, file4, file5, suff1, suff2, suff3, suff4, suff5, outfile):
    df1 = pd.read_csv(file1, sep="\t", dtype={'sample':'string'})
    df1 = df1.set_index('sample')
    df1 = df1.add_suffix("_"+suff1)
    df2 = pd.read_csv(file2, sep="\t", dtype={'sample':'string'})
    df2 = df2.set_index('sample')
    df2 = df2.add_suffix("_"+suff2)
    df3 = pd.read_csv(file3, sep="\t", dtype={'sample':'string'})
    df3 = df3.set_index('sample')
    df3 = df3.add_suffix("_"+suff3)
    df4 = pd.read_csv(file4, sep="\t", dtype={'sample':'string'})
    df4 = df4.set_index('sample')
    df4 = df4.add_suffix("_"+suff4)
    df5 = pd.read_csv(file5, sep="\t", dtype={'sample':'string'})
    df5 = df5.set_index('sample')
    df5 = df5.add_suffix("_"+suff5)

    result = pd.concat([df1, df2, df3, df4, df5], axis=1)
    result.to_csv(outfile,sep="\t")


def create_heatmap(infile1, infile2, outfile):
    df = pd.read_csv(infile1, sep="\t", dtype={'sample':'string'})
    df = df.set_index('sample')
    
    columns_split = []
    for col in df.columns:
        metric = col.split('_')[0]
        tool = '_'.join(col.split('_')[1:])
        columns_split.append((metric, tool, col))

    df.columns = pd.MultiIndex.from_tuples(
      [(metric, tool) for metric, tool, _ in columns_split], names=['Metric', 'Tool']
    )

    sorted_df = df.sort_index(axis=1, level='Metric')

    desired_tool_order = ['MiXCR', 'TRUST4_BAM', 'Vidjil', 'Conjunction', 'Disjunction'] 

    new_columns = []
    for metric in sorted_df.columns.get_level_values('Metric').unique():
        for tool in desired_tool_order:
            if (metric, tool) in sorted_df.columns:
                new_columns.append((metric, tool))

    heatmap_data = sorted_df.reindex(columns=new_columns)
    #heatmap_data = df.sort_index(axis=1, level=['Metric', 'Tool'])

    print(heatmap_data)
    
    df2 = pd.read_csv(infile2, sep="\t", dtype={'sample':'string'})
    df2 = df2.set_index('sample')
    data = df2.iloc[:, 4:6]
    
    result = pd.concat([data, heatmap_data], axis=1)
    print(result)
    
    #mask = data.isna()

    plt.figure(figsize=(22, 10))
    ax = sns.heatmap(result, annot=True, fmt='.2f', cmap="Reds", linewidths=0.5)
    ax.axvline(x=2, color='white', linewidth=10)
    ax.axvline(x=7, color='white', linewidth=10)
    ax.axvline(x=12, color='white', linewidth=10)
    ax.axvline(x=17, color='white', linewidth=10)
    ax.axvline(x=22, color='white', linewidth=10)
    ax.text(1, -0.15, 'Precision:\nMiXCR pipeline\nrelative to IGV', transform=ax.get_xaxis_transform(), 
        ha='center', va='bottom', fontsize=8, fontweight='bold')
    metrics = sorted(['Precision', 'Recall', 'F1', 'Specificity', 'Accuracy'])
    align_text_pos = [i for i in range(4, 28, 5)]
    for i in range(len(metrics)):
        ax.text(align_text_pos[i]+0.5, -0.22, metrics[i], transform=ax.get_xaxis_transform(), 
            ha='center', va='bottom', fontsize=8, fontweight='bold')
    ax.text(14.5, -0.27, 'TRUST4/MiXCR relative to confirmed in IGV clonotypes from MiXCR pipeline', transform=ax.get_xaxis_transform(),
        ha='center', va='bottom', fontsize=10, fontweight='bold')
    #plt.xlabel("Metric-tool", labelpad=60, fontweight='bold')
    plt.xlabel('')
    
    # remove y labels
    #plt.yticks(ticks=[], labels=[])

    plt.xticks(rotation=90, ticks=[i-0.5 for i in range(1, 28)], labels=['TCR', 'BCR']+['MiXCR', 'TRUST4_BAM', 'Vidjil', 'Conjunction', 'Disjunction']*5)
    plt.tight_layout()
    plt.savefig(outfile, dpi=800)


def scatter_plot(infile1, infile2, outfile):

    df_cov = pd.read_csv(infile2, sep="\t", dtype={'sample':'string'})
    df_cov = df_cov[["sample", "Median Coverage"]]
    df_cov = df_cov.set_index('sample')

    df = pd.read_csv(infile1, sep="\t", dtype={'sample':'string'})
    df = df.set_index('sample')
    
    df2 = pd.concat([df, df_cov], axis=1)
    
    fig, axes = plt.subplots(2, 3, figsize=(18, 8))
    fig.suptitle("The metric's dependence on coverage", fontsize=16, y=1)
    
    metrics = ['Precision', 'Recall', 'F1', 'Specificity', 'Accuracy']
    fig_pos_x = [0, 0, 0, 1, 1]
    fig_pos_y = [0, 1, 2, 0, 1]
    
    for i in range(0, 5):
        metric = metrics[i]
        df_long = pd.concat([
            df2[['Median Coverage',metric+'_TRUST4_BAM']].assign(Program='TRUST4_BAM').rename(columns={metric+'_TRUST4_BAM': metric}),
            df2[['Median Coverage', metric+'_MiXCR']].assign(Program='MiXCR').rename(columns={metric+'_MiXCR': metric}),
            df2[['Median Coverage', metric+'_Vidjil']].assign(Program='Vidjil').rename(columns={metric+'_Vidjil': metric}),
            df2[['Median Coverage', metric+'_Conjunction']].assign(Program='Conjunction').rename(columns={metric+'_Conjunction': metric}),
            df2[['Median Coverage', metric+'_Disjunction']].assign(Program='Disjunction').rename(columns={metric+'_Disjunction': metric})
        ], ignore_index=True)

        sns.scatterplot(
          data=df_long,
          x='Median Coverage',
          y=metric,
          hue='Program',
          palette={'MiXCR': '#A90606', 'TRUST4_BAM': '#FA9F42', 'Vidjil':"#2B4162", 'Conjunction':"#0B6E4F", 'Disjunction':"#59C3C3"},
          s=60,
          ax=axes[fig_pos_x[i], fig_pos_y[i]]
        )

        sns.regplot(data=df2, x='Median Coverage', y=metric+'_TRUST4_BAM', scatter=False, color='#FA9F42', ax=axes[fig_pos_x[i], fig_pos_y[i]])
        sns.regplot(data=df2, x='Median Coverage', y=metric+'_MiXCR', scatter=False, color='#A90606', ax=axes[fig_pos_x[i], fig_pos_y[i]])
        sns.regplot(data=df2, x='Median Coverage', y=metric+'_Vidjil', scatter=False, color='#2B4162', ax=axes[fig_pos_x[i], fig_pos_y[i]])
        sns.regplot(data=df2, x='Median Coverage', y=metric+'_Conjunction', scatter=False, color='#0B6E4F', ax=axes[fig_pos_x[i], fig_pos_y[i]])
        sns.regplot(data=df2, x='Median Coverage', y=metric+'_Disjunction', scatter=False, color='#59C3C3', ax=axes[fig_pos_x[i], fig_pos_y[i]])
        
        #axes[fig_pos_x[i], fig_pos_y[i]].set_title(metric)
        axes[fig_pos_x[i], fig_pos_y[i]].set_ylabel(metric)

    axes[1, 2].set_visible(False)
    #plt.xlabel('Median Coverage')
    #plt.ylabel('Precision')
    #plt.legend(title='Tool')
    plt.tight_layout()
    plt.grid(True, alpha=0.3)
    plt.savefig(outfile, dpi=800)


def summary_stat(infile, outfile):
    df = pd.read_csv(infile, sep="\t")
    pd.set_option('display.float_format', '{:.2f}'.format)
    means = df.mean()*100
    
    parsed_data = []
    for line in means.reset_index().values.tolist()[1:]:
        metric_tool = line[0]  
        value = float(line[1])

        metric, tool = metric_tool.split('_', 1)
        parsed_data.append({'tool': tool, 'metric': metric, 'value': value})


    df_parsed = pd.DataFrame(parsed_data)

    df_table = df_parsed.pivot(index='tool', columns='metric', values='value')
    print(df_table)
    df_table.to_csv(outfile, sep="\t")



create_stat_table("samples_clones.xlsx", "mixcr_stat.csv", 8)
create_stat_table("samples_clones.xlsx", "trust4_bam_stat.csv", 6)
create_stat_table("samples_clones.xlsx", "vidjil_stat.csv", 9)
create_stat_table("samples_clones.xlsx", "AND_trust4_mixcr_vidjil_stat.csv", 10)
create_stat_table("samples_clones.xlsx", "OR_trust4_mixcr_vidjil_stat.csv", 11)
create_stat_table_bcr_tcr("samples_clones.xlsx", "samples_coverage_copy.csv", "mixcr_igv_bcr_tcr.csv")
concat_tables("trust4_bam_stat.csv", "mixcr_stat.csv", "vidjil_stat.csv", "AND_trust4_mixcr_vidjil_stat.csv", "OR_trust4_mixcr_vidjil_stat.csv", 'TRUST4_BAM', 'MiXCR', 'Vidjil', 'Conjunction', 'Disjunction', "mixcr_igv_statistics.csv")

# plot
create_heatmap("mixcr_igv_statistics.csv", "mixcr_igv_bcr_tcr.csv", "heatmap_metrics.png")

#scatter_plot("mixcr_igv_statistics.csv", "mixcr_igv_bcr_tcr.csv", "scatterplot.png")

#summary_stat("mixcr_igv_statistics.csv", "summary.txt")
